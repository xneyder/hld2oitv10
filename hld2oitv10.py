#!/usr/bin/python
""" hld2oit.py:

 Description: Tool intended to convert HLD format files to OIT format


 Input Parameters:
		HLD File: Location to the HLD excel file
		Custom counter file: File containing a json file defining counters with
            custom formulas

 Output: OIT excel file

 Example:
		hld2oit.py "HLD_USC_AFF_vMCC_V.1.0.2.xls"

 Database:	N/A

 Created by : Daniel Jaramillo
 Creation Date: 04/01/2019
 Modified by:     Date:
 All rights(C) reserved to Teoco
"""

import sys
import os
import tokenize
import ast
import json
import copy
import pandas as pd
import numpy as np
from StringIO import StringIO
from openpyxl import load_workbook
from LoggerInit import LoggerInit
from oit_mapping import oit_mapping

def get_vars_divs(formula):
    """
    Description: Gets the variables and the divisors from a formula
    Input Parametes:
        formula
    """
    vars=[]
    divs=[]
    div=''
    start_div=False
    it=tokenize.generate_tokens(StringIO(formula).readline)
    for type,value,_,_,_ in it:
        if value=='/':
            if div:
                divs.append(div)
                div=''
            start_div=True
        elif value=='*':
            if div:
                divs.append(div)
                div=''
            start_div=False
        elif value==')':
            if div:
                divs.append(div)
                div=''
            start_div=False
        else:
            if type == 1:
                vars.append(value)
            if start_div and value !="(":
                div+=value
    if div:
        divs.append(div)
    vars=list(set(vars))
    divs=list(set(divs))
    return vars,divs

def create_tpt(kpi_name,formula,folder,table):
    """
    Description: Creates tpt function fomr a formula
    Input Parametes:
       kpi_name
       formula
       folder
       table
    """
    app_logger=logger.get_logger("create_tpt")
    app_logger.info("Creating {kpi_name}={formula}"\
                    .format(kpi_name=kpi_name,formula=formula))
    vars,divs=get_vars_divs(formula)
    schema=metadata['Library Info']['VENDOR']\
            +"_"+metadata['Library Info']['DOMAIN']
    #Build the string to call the function
    call_str='{schema}_{kpi_name}('.format(schema=schema,kpi_name=kpi_name)
    call_vars=[]
    for var in vars:
        if var+table in temp_dict:
            rd_name=temp_dict[var+table]
        else:
            index=metadata['Counters_KPI']\
               .index[(metadata['Counters_KPI']['Counter/KPI DB Name']\
                      == var)
                      & (metadata['Counters_KPI']['Table Name']\
                      == table)]
            rd_name='{'+metadata['Counters_KPI']\
               .loc[index,'Raw Data Counter Name/OID'].item()+'}'
        call_vars.append('{rd_name}'.format(rd_name=rd_name))
    call_str+=','.join(call_vars)+')'
    #Check if KPI is already created
    id='{schema}_{kpi_name}\n'.format(schema=schema,kpi_name=kpi_name)
    if id in tpt_functions:
        return call_str
    tpt_functions.append(id)
    tpt_file_name='{schema}_TrolLocalFunctions.tpt'\
        .format(schema=schema)
    with open(tpt_file_name,'a') as file:
        file.write('\n')
        file.write('@@PROTO\n')
        file.write('type=UF\n')
        file.write('id={schema}_{kpi_name}\n'\
                   .format(schema=schema,kpi_name=kpi_name))
        file.write('location=Local.{folder}\n'.format(folder=folder))
        file.write('desc=\n')
        file.write('bitmap=\n')
        file.write('inpParamsNum={num_vars}\n'.format(num_vars=len(vars)))
        for idx,var in enumerate(vars):
            file.write('{idx}={var}, double, 1\n'.format(idx=idx+1,var=var))
        file.write('outParamsNum=1\n')
        file.write('1={kpi_name}, double\n'.format(kpi_name=kpi_name))
        file.write('keywordsNum=0\n')
        file.write('help=\n')
        file.write('@@CodeBegin\n')
        file.write('\n')
        for idx,var in enumerate(vars):
            vars[idx]='IsNull({var})'.format(var=var)
        vars_val="if ({vars}){{".format(vars='||'.join(vars))
        file.write(vars_val)
        file.write('\n')
        file.write('    {kpi_name} = NullDouble();\n'.format(kpi_name=kpi_name))
        file.write('    return true;\n')
        file.write('}\n')
        if divs:
            for idx,div in enumerate(divs):
                divs[idx]='{div} == 0'.format(div=div)
            divs_val="if ({divs}){{".format(divs='||'.join(divs))
            file.write(divs_val)
            file.write('\n')
            file.write('    {kpi_name} = 0;\n'.format(kpi_name=kpi_name))
            file.write('    return true;\n')
            file.write('}\n')
        file.write('{kpi_name}={formula};'\
                   .format(kpi_name=kpi_name,formula=formula))
        file.write('\n')
        file.write('return true;\n')
        file.write('\n')
        file.write('@@CodeEnd\n')
        file.write('@@PROTO_END\n')
        file.write('\n')
    return call_str


def create_functions():
    """
    Description: creates the functions for the KPI counters
    """
    global custom_counters
    global temp_dict
    app_logger=logger.get_logger("create_functions")
    app_logger.info("Creating functions")
    schema=metadata['Library Info']['VENDOR']\
            +"_"+metadata['Library Info']['DOMAIN']
    tpt_file_name='{schema}_TrolLocalFunctions.tpt'\
        .format(schema=schema)
    #Make file emty
    open(tpt_file_name, 'w').close()
    temp_cnt=1
    #Loop over all counters
    df=metadata['Counters_KPI']
    for idx,kpi in df.iterrows():
        raw_formula=kpi['KPI Formula']
        if kpi['Counter/KPI DB Name'] in custom_counters:
            call_str=custom_counters[kpi['Counter/KPI DB Name']]['call_str']
            if custom_counters[kpi['Counter/KPI DB Name']]['generate_temp']\
               == 'True':
                #Generate temp entry
                temp_dict[kpi['Counter/KPI DB Name']+kpi['Table Name']]\
                        ='temp{temp_cnt}'.format(temp_cnt=temp_cnt)
                temp_cnt+=1
                pass
        elif isinstance(raw_formula, float) and np.isnan(raw_formula):
            continue
        else:
            formula=kpi['KPI Formula'].replace(' ','').replace('\n','')
            #Validate that the formula is valid
            try:
                ast.parse(formula)
            except SyntaxError:
                app_logger.error('Wrong formula {kpi_name}={formula}'\
                                 .format(kpi_name=kpi['Counter/KPI DB Name'],
                                         formula=formula))
                quit()

            call_str=create_tpt(kpi['Counter/KPI DB Name'],
                formula,
                schema,
                kpi['Table Name'])
        #Modify formula in metadata
        index=metadata['Counters_KPI']\
                .index[(metadata['Counters_KPI']['Counter/KPI DB Name']\
                       ==kpi['Counter/KPI DB Name'])\
                      & (metadata['Counters_KPI']['Table Name']\
                       ==kpi['Table Name'])]
        metadata['Counters_KPI'].loc[index,'KPI Formula']=call_str

def parse_front_page(xl):
    """
    Description:  Parse the Front Page sheet
    Input Parametes:
        xl: Pandas excel file object
    """
    global metadata
    metadata['Front Page']={}
    app_logger=logger.get_logger("parse_front_page")
    app_logger.info("Parsing front page")
    df=xl.parse('Front Page')
    df=df.iloc[:,[0,1]].dropna(how='all')
    for index,row in df.iterrows():
        if row[0] == "Revision History":
            break
        metadata['Front Page'][row[0]]=row[1]

def parse_library_info(xl):
    """
    Description:  Parse the Library Info sheet
    Input Parametes:
        xl: Pandas excel file object
    """
    global metadata
    metadata['Library Info']={}
    app_logger=logger.get_logger("parse_library_info")
    app_logger.info("Parsing Library Info")
    df=xl.parse('Library Info')
    #df=df.iloc[:,[1,2]].dropna(how='all')
    df=df.iloc[:,[1,2]]
    for index,row in df.iterrows():
        try:
            if index[1] == "Table Retention:":
                break
            metadata['Library Info'][index[1]]=index[2]
        except IndexError:
            continue
    metadata['Library Info']['SCHEMA']=metadata['Library Info']['VENDOR']\
            +"_"+metadata['Library Info']['DOMAIN']



def parse_table(xl,sheet_name):
    """
    Description:  Parse the sheet in table format
    Input Parametes:
        xl: Pandas excel file object
        sheet name
    """
    global metadata
    metadata[sheet_name]={}
    app_logger=logger.get_logger("parse_table")
    app_logger.info("Parsing {sheet_name}".format(sheet_name=sheet_name))
    df=xl.parse(sheet_name)
    metadata[sheet_name]=df.iloc[2:,1:]


def load_hld(hld_file):
    """
    Description: Load the configuration from HLD file
    Input Parametes:
        hld_file: Excel containing the functional specification for the library
    """
    app_logger=logger.get_logger("load_hld "+hld_file)
    app_logger.info("Parsing HLD")
    xl=pd.ExcelFile(hld_file)
    parse_front_page(xl)
    parse_library_info(xl)
    parse_table(xl,"Entities")
    parse_table(xl,"Tables")
    parse_table(xl,"Counters_KPI")



def write_oit():
    """
    Description: write to OIT
    Input Parametes:
        hld_file: Excel containing the functional specification for the library
    """
    app_logger=logger.get_logger("write_oit")
    app_logger.info("Creating OIT File")

    wb = load_workbook('template/EASY_PM_TEMPLATE_HELIX10.xlsx')

    #Populate Front Page
    for sheet,fields in oit_mapping.items():
        ws = wb[sheet]
        for field in fields:
            value=metadata[field['hld_sheet']][field['hld_field']]
            ws.cell(row=field['row'], column=field['column'], value=value)

    #Populate Entities related sheets
    ws_ent = wb['Entities']
    ws_cfg_fields = wb['Configuration Fields']
    for index,entity in metadata['Entities'].iterrows():
        #Populate Entities
        schema=entity['CFG Table or conf View'].split('.')[0]
        configuration_view=entity['CFG Table or conf View'].split('.')[1]
        #Get table list for autopuplate
        df=metadata['Tables']
        df=df.loc[df['Entity'] == entity['Entity Name']].head(3)
        tables_arr=[]
        for index,table in df.iterrows():
            tables_arr.append(table['Table Name'])
        tables=','.join(tables_arr)

        record=[entity['Entity Name'],
                entity['Entity Type'],
                entity['Display Name'],
                entity['Element Alias'],
                entity['Parent Entity'],
                '', #Domain
                schema,
                entity['Presentation'],
                configuration_view,
                entity['Universe'],
                entity['BC Object Type '],
                'Y',
                'N',
                tables,
                'N',
                5,
               ]

        ws_ent.append(record)
        #Populate CFG Fields
        key_list=entity['Keys'].split(',')
        for idx,key in enumerate(key_list):
            record=[configuration_view,
                    key,
                    'VARCHAR2',
                    'Y',
                    100,
                    idx+1]
            ws_cfg_fields.append(record)

    #Populate Counter Sets
    ws_cs = wb['Counter Sets']
    for index,table in metadata['Tables'].dropna(how='all').iterrows():
        #Fill Counter Sets
        summaries=table['Time Summary'].split(',')
        _5M=''
        _15M=''
        _30M=''
        HR=''
        DY=''
        WK=''
        MO=''
        YR=''
        if '5M' in summaries:
            _5M='1M'
        if '15M' in summaries:
            _15M='5M'
        if '30M' in summaries:
            _30M='15M'
        if 'HR' in summaries:
            HR='15M'
        if 'DY' in summaries:
            DY='HR'
        if 'WK' in summaries:
            WK='DY'
        if 'MO' in summaries:
            MO='DY'
        if 'YR' in summaries:
            YR='MO'
        record=[table['Table Name'],
                table['Counter Group Display Name'],
                table['Alias Table Name '],
                table['Counter Group in RD'],
                table['Entity'],
                'Y',
                '',
                table['Universe'],
                table['Base Granularity'],
                _5M,
                _15M,
                _30M,
                HR,
                DY,
                WK,
                MO,
                YR,
               ]
        ws_cs.append(record)

    #Populate Loaded Counters
    ws = wb['Loaded Counters']
    df=metadata['Counters_KPI'].dropna(how='all')
    aggr_list=['AVG','SUM','MAX','MIN']
    temp_ct=1
    order={}
    for index,counter in df.iterrows():
        size=''
        if counter['TYPE'] in ['GPI','PI','OI']:
            size=100
        if counter['Hierarchy Summary'] not in aggr_list:
            ent_aggr_formula='NULL'
        else:
            ent_aggr_formula=counter['Hierarchy Summary']
        record=[counter['Table Name'],
                counter['Counter/KPI DB Name'],
                counter['Counter/KPI Display Name'],
                counter['Vendor Counter Name'],
                '',
                counter['TYPE'],
                counter['KPI Formula'],
                size,
                counter['Counter Description'],
                '',
                'Y',
                'Y',
                '',
                counter['15M'],
                counter['30M'],
                counter['HR'],
                counter['DY'],
                counter['WK'],
                counter['MO'],
                counter['YR'],
                ent_aggr_formula,
                ent_aggr_formula,
                ent_aggr_formula,
        ]
        #Counter has custom formula and is needed a temp counter
        if counter['Counter/KPI DB Name'] in custom_counters and\
           custom_counters[counter['Counter/KPI DB Name']]['generate_temp']\
               == 'True':
            temp_record=copy.deepcopy(record)
            temp_record[1]='temp{temp_ct}'.format(temp_ct=temp_ct)
            #Fix the formula to use the temp counter
            temp_ct+=1
            temp_record[2]=''
            temp_record[3]=''
            temp_record[5]='NULL'
            temp_record[6]=counter['KPI Formula']
            #Increase the order of the counter
            temp_record[10]='N'
            temp_record[11]='N'
            temp_record[13]='NULL'
            temp_record[14]='NULL'
            temp_record[15]='NULL'
            temp_record[16]='NULL'
            temp_record[17]='NULL'
            temp_record[18]='NULL'
            temp_record[19]='NULL'
            temp_record[20]='NULL'
            temp_record[21]='NULL'
            temp_record[22]='NULL'
            ws.append(temp_record)
        ws.append(record)
    schema=metadata['Library Info']['VENDOR']\
            +"_"+metadata['Library Info']['DOMAIN']
    wb.save("{schema}_EZPM.xlsx".format(schema=schema))
    app_logger.info("{schema}_EZPM.xlsx file created".format(schema=schema))

def main():
    global custom_counters
    app_logger=logger.get_logger("main")
    app_logger.info("Starting {script}".format(script=sys.argv[0]))
    #Validate the line arguments
    if len(sys.argv) < 3:
        app_logger.error("Usage {script} <HLD File> [custom counter file]"
                         .format(script=sys.argv[0]))
        app_logger.error("Example {script} 'HLD_USC_AFF_vMCC_V.1.0.2.xls'\
                          AFFIRMED_vMCC_custom_counters.py"
                         .format(script=sys.argv[0]))
        quit()
    hld_file=sys.argv[1]
    #Are there custom counters?
    if sys.argv[2]:
        custom_counters_file=sys.argv[2]
        try:
            with open(custom_counters_file) as json_file:
                custom_counters=json.load(json_file)
        except IOError:
            pass
    #Load configuration
    load_hld(hld_file)
    #Create tpt functions
    create_functions()
    #Create OIT
    write_oit()


#Application starts running here
if __name__ == "__main__":
    #If LOG_DIR environment var is not defined use /tmp as logdir
    if 'LOG_DIR' in os.environ:
        log_dir=os.environ['LOG_DIR']
    else:
        log_dir="/tmp"

    log_file=os.path.join(log_dir,"hld2oit.log")
    logger=LoggerInit(log_file,10)
    metadata={}
    custom_counters={}
    temp_dict={}
    tpt_functions=[]
    main()
